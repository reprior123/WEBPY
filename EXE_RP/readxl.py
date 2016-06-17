
##from openpyxl import load_workbook
##wb = load_workbook(filename = r'ASSETS by revenue.xlsx')
##print wb.get_sheet_names()
##sheet_ranges = wb['range names']
##print sheet_ranges['D18'].value # D18

##import xlrd
##import csv
##
##with xlrd.open_workbook('a_file.xls') as wb:
##    sh = wb.sheet_by_index(0)  # or wb.sheet_by_name('name_of_the_sheet_here')
##    with open('a_file.csv', 'wb') as f:
##        c = csv.writer(f)
##        for r in range(sh.nrows):
##            c.writerow(sh.row_values(r))


## for xlsx files...           
import openpyxl, csv

firm = 'agz'

for firm in ['agz', 'inc', 'ltd']:
    infile = firm +'.sum.xlsx'
    outfile =  firm +'.sum.csv'
    wb = openpyxl.load_workbook(infile)
    sh = wb.get_active_sheet()
    with open(outfile, 'wb') as f:
        c = csv.writer(f)
        for r in sh.rows:
            c.writerow([cell.value for cell in r])
